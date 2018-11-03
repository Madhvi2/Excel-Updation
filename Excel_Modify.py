# -*- coding: utf-8 -*-
"""
Created on Wed Oct 31 18:37:09 2018

@author: m37
"""

from string import ascii_lowercase
from openpyxl import load_workbook

def mapping(col_name):
    LETTERS = {letter: str(index) for index, letter in enumerate(ascii_lowercase, start=1)}

    col_name = col_name.lower()
    for key,value in LETTERS.items():
        if(key==col_name):
            index = LETTERS[key]
    return (index)

def readCandidateFile(file_path,sheet_no,index):
    wb = load_workbook(file_path)
    ws = wb[sheet_no]
    data = {}
    for row in ws: 
        if ws.row_dimensions[row[0].row].hidden == False:
            for cell in row:
                col = int(mapping(cell.column))
                if(col == 1):
                    key = cell.value
                if(col == int(index)):
                    data[key]=cell.value
    del(data[list(data.keys())[0]])
    return data

def readMainFile(file_path,sheet_no,index):
    wb = load_workbook(file_path)
    ws = wb[sheet_no]
    data = {}
    index = int(index)
    for i in range(1,ws.max_row+1):
        data[ws.cell(row=i,column=1).value] = ws.cell(row= i,column=index).value
    return data
    
def compareFiles(data_dict_1,data_dict_2):
    for k in data_dict_2.keys():
        data_dict_1[k] = data_dict_2[k]
    return(data_dict_1)   
 
def writeValues(file_path,sheet_no,updated_list,index):
    wb = load_workbook(file_path)
    ws = wb[sheet_no]
    index= int(index)
    for i in range(0,ws.max_row):
        print(i)
        ws.cell(row=i+1,column=index).value = updated_list[i]
    wb.save(file_path)
    
def main():
    #file_path_1 = input("Enter file path to be updated")
    #sheet_no_1 = 'Sheet'+input("Enter the sheet name to be updated")
    #col_name_1 = input("Enter the column name to be updated")
    #file_path_2 = input("Enter file path from which value need to be updated")
    #sheet_no_2 = 'Sheet'+input("Enter the sheet name from which value need to be updated")
    #col_name_2 = input("Enter the column name from which value need to be updated")
    
    file_path_2 = "C:/Users/m37/Excel Automation/1.xlsx"
    sheet_no_2 = "Sheet1"
    col_name_2 = "B"
    
    file_path_1 = "C:/Users/m37/Excel Automation/Main.xlsx"
    sheet_no_1 = "Sheet1"
    col_name_1 = "B"
    
    index_1 = mapping(col_name_1)
    index_2 = mapping(col_name_2)
    
    data_dict_1 = readMainFile(file_path_1,sheet_no_1,index_1)
    data_dict_2 = readCandidateFile(file_path_2,sheet_no_2,index_2)
    updated_list = compareFiles(data_dict_1,data_dict_2)
    updated_list = list(updated_list.values())
    writeValues(file_path_1,sheet_no_1,updated_list,index_1)

if __name__== "__main__":
    main()